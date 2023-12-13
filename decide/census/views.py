from django.db.utils import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from django.views import View
from django.views.generic.base import TemplateView
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.contrib import messages
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.status import (
    HTTP_201_CREATED as ST_201,
    HTTP_204_NO_CONTENT as ST_204,
    HTTP_400_BAD_REQUEST as ST_400,
    HTTP_401_UNAUTHORIZED as ST_401,
    HTTP_409_CONFLICT as ST_409,
)

from base.perms import UserIsStaff
from .forms import CreationCensusForm
from .models import Census
from voting.models import Voting
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from django.shortcuts import redirect
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User

def census(request):
    return render(request,'census/census.html')

class CensusCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = Census(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response("Error try to create census", status=ST_409)
        return Response("Census created", status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = Census.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusDetail(generics.RetrieveDestroyAPIView):
    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get("voters")
        census = Census.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response("Voters deleted from census", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get("voter_id")
        try:
            Census.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")
    
def GetId(request):
    id = request.GET['id']
    
    census = Census.objects.filter(voting_id=int(id))
    if len(census) == 0:
        return render(request,'census/census.html',{'error_id':'There is not a census with that voting_id'})
    else:
        return render(request,"census/census_details.html",{'census':census})


class CensusExportView(TemplateView):
    template_name = "census/export_census.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        votings = Voting.objects.all()
        context["votings"] = votings
        return context


def export_census(request, voting_id):
    voting = Voting.objects.get(id=voting_id)
    census = Census.objects.filter(voting_id=voting_id)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=census.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Census"

    worksheet.merge_cells("A1:B1")
    first_cell = worksheet["A1"]
    first_cell.value = "Census for: " + voting.name
    first_cell.fill = PatternFill("solid", fgColor="0D98BA")
    first_cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
    first_cell.alignment = Alignment(horizontal="center")

    voting_id_title = worksheet["A2"]
    voting_id_title.value = "Voting ID"
    voting_id_title.font = Font(bold=True, color="0D98BA", name="Calibri")
    voting_id_title.alignment = Alignment(horizontal="center")

    voter_id_title = worksheet["B2"]
    voter_id_title.value = "Voter ID"
    voter_id_title.font = Font(bold=True, color="0D98BA", name="Calibri")
    voter_id_title.alignment = Alignment(horizontal="center")

    for row in census:
        row_data = [row.voting_id, row.voter_id]
        worksheet.append(row_data)

    workbook.save(response)

    return response


class CensusImportView(TemplateView):
    template_name = "census/import_census.html"

    def post(self, request, *args, **kwargs):
        if request.method == "POST" and request.FILES:
            file = request.FILES["file"]
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                voting_id = row[0]
                voter_id = row[1]

                try:
                    Census.objects.create(voting_id=voting_id, voter_id=voter_id)
                except Exception as e:
                    messages.error(request, f"Error importing data: {str(e)}")
                    return HttpResponseRedirect("/census/import/")

            messages.success(request, "Data imported successfully!")
            return HttpResponseRedirect("/census/import/")
        if request.method == "POST" and not request.FILES:
            messages.error(request, "No file selected!")
            return HttpResponseRedirect("/census/import/")

######Creación de censo

def createCensus(request):
    if request.method == 'POST':
        voting_id = request.POST.get('voting_id')
        voter_id = request.POST.get('voter_id')

        try:
            census = Census.objects.create(voting_id=voting_id, voter_id=voter_id)
            census.full_clean()
            census.save()
            messages.success(request, 'Census created successfully')
            return redirect('census')
        
        # Si hay un ValidationError, muestra el mensaje de error en la página de creación del censo
        except ValidationError as e:
            if not Voting.objects.filter(id=voting_id).exists():
                return render(request, 'census/census_create.html', {'error':'Voting with this ID does not exist.', 'form': CreationCensusForm})

            if not User.objects.filter(id=voter_id).exists():
                return render(request, 'census/census_create.html', {'error': 'User with this ID does not exist.', 'form': CreationCensusForm})
            
    # Si el método no es POST, muestra la página de creación del censo
    return render(request, 'census/census_create.html',{'form': CreationCensusForm})


############BORRAR CENSOS
def deleteCensus(request):
    census = Census.objects.filter(voting_id=request.POST['voting_id'], voter_id = request.POST['voter_id'])
    if len(census) == 0: 
        return render(request,'census/census.html',{'error':'Census does not exist. Try other census'})
    if len(census) != 0:
        census.delete()
        messages.success(request, 'Census deleted successfully')
        return redirect('census')

def censusList(request):
    queryset = Census.objects.all()
    return render(request, 'census/census_list.html', {'queryset':queryset})

